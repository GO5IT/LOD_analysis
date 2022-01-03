from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout

# Save as EXCEL for the last dafatrame (all merged one) in allagg aray
def save_excel(allagg, coverage_df, fileprefix, entityname):
    wb = Workbook()
    # Access the default sheet and rename it
    ws = wb.active
    ws.title = 'comparison'
    #Convert dataframe to rows in EXCEL and add them to the sheet
    for r in dataframe_to_rows(allagg[-1], index=True, header=True):
       ws.append(r)
    # Alternative way to save as EXCEL or CSV
    #allagg[-1].to_excel(r'allagg[1].xlsx', sheet_name='comparison', index = False)
    #allagg_csv = allagg[-1].to_csv('allagg[1].csv', index=True, index_label='@id')

    # Create a sheet at the end of the workbook
    ws1 = wb.create_sheet('coverage')
    # Convert dataframe to rows in EXCEL and add them to the sheet
    #allagg_csv = coverage_df.to_csv('000.csv', index=True, index_label='@id')
    #for r in dataframe_to_rows(coverage_df, index=False, header=False):
    #   ws1.append(r)

    # Create sheet from dataframe (iterate each rows in EXCEL)
    for r in dataframe_to_rows(coverage_df, index=False, header=True):
        ws1.append(r)
    # print(ws1['B3'].value) D3
    # print(ws1['D11'].value) F5

    chart1 = BarChart()
    label = Reference(ws1, min_col=1, min_row=2, max_row=12)
    data_range = Reference(ws1, min_col=4, max_col=5, min_row=1, max_row=12)
    chart1.add_data(data_range, titles_from_data=True)
    chart1.set_categories(label)
    # chart1.shape = 4
    # chart1.type = 'col'
    chart1.title = 'Bar Chart'
    chart1.y_axis.title = 'Coverage'
    chart1.x_axis.title = 'Source'
    chart1.type = 'bar'
    chart1.style = 13
    chart1.grouping = 'percentStacked'
    chart1.overlap = 100
    chart1.title = 'Content Coverage of Sources'
    chart1.layout = Layout(
        manualLayout=ManualLayout(
            x=0.25, y=0.25,
            h=1.0, w=1.0,
        )
    )
    # Specify the starting point of chart
    ws1.add_chart(chart1, 'G1')

    # Save Workbook (2 sheets) as XLSX. Dates should be string instead of integer (entityname)
    filename = fileprefix + str(entityname) + '.xlsx'
    # Use below to save in a testing file
    #filename = '00001.xlsx'
    wb.save(filename)
    return()