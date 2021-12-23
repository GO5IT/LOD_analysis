import pickle
import rdfpandas
from rdfpandas.graph import to_dataframe
import pandas as pd
import rdflib
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Option a: include all access URLs and URLs manually
# Wikidata generates an error with "rdf:nodeID value is not a valid NCName"
# Create an array for entity URIs
# data = {
# 'source' : ['Worldcat', 'LoC', 'VIAF', 'Getty',	'Wikidata', 'DBpedia', 'BabelNet', 'GeoNames', 'YAGO', 'Europeana'],
# # Below is the list of entities to be accessed (with file extention)
# 'url' : ['http://id.worldcat.org/fast/1037825.rdf.xml',
# 'https://id.loc.gov/authorities/subjects/sh97003090.rdf',
# '',
# '',
# #'http://www.wikidata.org/entity/Q2648',
# #'https://www.wikidata.org/wiki/Special:EntityData/Q2648.rdf',
# '',
# 'https://dbpedia.org/data/1967.rdf',
# 'http://babelnet.org/rdf/data/s02830721n?output=xml',
# '',
# '',
# ''],
# # Below is the list of URIs to be searched (without file extention except BabelNet. Watchout also https)
# 'uri' : ['http://id.worldcat.org/fast/1037825',
# 'http://id.loc.gov/authorities/subjects/sh97003090',
# '',
# '',
# #'http://www.wikidata.org/entity/Q2648',
# '',
# 'http://dbpedia.org/resource/1967',
# 'http://babelnet.org/rdf/data/s02830721n?output=xml',
# '',
# '',
# '']
# }
# entities = pd.DataFrame(data)
# print(entities)

# Option b: include all access URLs and URLs automatically from EXCEL file
xl_file = pd.ExcelFile("ListOfAllURLs_nohttps_for_URIs_and_URLs_GettyWebURIs_BabelNameSpaces_GettyNameSpaces_wikidataNameSpaces.xlsx")
# Put 2 sheets in DataFrame with the sheetnames
data = {sheet_name: xl_file.parse(sheet_name)
           for sheet_name in xl_file.sheet_names}
#print(entities['ListOfURLs']['YAGO'])
#print(entities['ListOfURIs'].T.loc["Entity"][0])
# Transpose to make a right shape of Dataframe and locate the series (e.g. series for Shakespeare)
# Iterate over all entities and generate CSV separately (ca 3000), and for each category (ca 60,000) all is too long (300,000)
url = data['ListOfURLs'].T.loc[:,7]
uri = data['ListOfURIs'].T.loc[:,7]
# Merge two series from two sheets
merged = pd.concat([url, uri], axis=1)
# Replace NaN with empty cell
merged_null = merged.fillna("")
# Dataframe cleanup by dropping unneeded rows, resetting index, naming columns
entities = merged_null.drop(merged_null.index[0:2])
entities.reset_index(inplace=True)
entities.columns = ['source', 'url', 'uri']
print(entities)

# If some URLs do not work, you can ignore them, by replacing problematic URLs and URLs with empty (e.g. Wikidata and Europeana)
# entities.loc[3:5, ['url', 'uri']] = ''
#entities.loc[4, ['url', 'uri']] = ''
entities.loc[9, ['url', 'uri']] = ''
print(entities)

# # Option A: Check each entity manually
# # Set which row to analyse by providing row number in [] as entities to compare (to find the most important data in an entity)
# rownumber = 6
# target = entities.loc[rownumber, "url"]
# target2 = entities.loc[rownumber, "uri"]
# name = entities.loc[rownumber, "source"]
# # If dataframe is not empty ("")
# if target != "":
#     print('\n\n***********************\nWorking on ' + name + '\n***********************\n\n')
#     g = rdflib.Graph()
#     g.parse(target, format = 'xml')
#     df = to_dataframe(g)
#     df['entity'] = df.index
#     print(df)
#     # Exact match
#     x = df[df['entity'] == target2]
#     #Test: Get rows that contain specified text (in this case a string from dataframe)(not used)
#     #x = df.loc[df["entity"].str.contains(target2, regex=False)]
#     x_transposed = x.T
#     print(x_transposed)
#     # #Saving as CSV file
#     df_csv = x_transposed.to_csv('Manual_Entity_Check.csv', index=True, index_label="@id")
# else:
#     print('\n\n***********************\n' + name + ' does not have entity. Try another entity!!\n***********************\n\n')

# Option B: Create dataframe for each entity automatically
aggregate = []
# create dataframe for each entity
for i in range(len(entities)):
    print('\n\n***********************\nWorking on ' + entities.loc[i, "source"] + '\n***********************\n\n')
    if entities.loc[i, "url"] != "":
        g = rdflib.Graph()
        # Wikidata works only with ttl
        if entities.loc[i, "source"] == "Wikidata":
            g.parse(entities.loc[i, "url"], format = 'ttl')
        else:
            g.parse(entities.loc[i, "url"], format = 'xml')
        df = to_dataframe(g)
        df['entity'] = df.index
        #print(df.head)
        # Check matching rows in the dafafame which includes the entity URI
        x = df[df['entity'] == entities.loc[i, "uri"]]
        x_transposed = x.T
        x_transposed['full_coverage'] = x_transposed[entities.loc[i, "uri"]]
        x_transposed.dropna(subset=['full_coverage'], inplace=True)
        print(x_transposed)
        aggregate.append(x_transposed)
        print(aggregate)
        # Save dataframe of each entity as CSV (file name is index number (i) in the dataframe)
        #csvnumber = entities.loc[i, "source"] + str(i) + '.csv'
        #df_csv = x_transposed.to_csv(csvnumber, index=True, index_label="@id")
    else:
        print('\n\n***********************\n' + entities.loc[i, "source"] + ' does not have entity. Try another entity!!\n***********************\n\n')

# #Option X: Check 2 dataframes manually
# #print(aggregate[1])
# aggregate[0].dropna(subset = ['full_coverage'], inplace=True)
# aggregate[1].dropna(subset = ['full_coverage'], inplace=True)
# # Save as CSV
# agg1 = aggregate[0].to_csv('Manual_BeforeMerge_world.csv', index=True, index_label="@id")
# agg2 = aggregate[1].to_csv('Manual_BeforeMerge_loc.csv', index=True, index_label="@id")
# merged = aggregate[0].merge(aggregate[1].drop_duplicates(), on='full_coverage', how='outer', indicator=True, validate='many_to_many')
# #print(merged)
# merged_csv = merged.to_csv('Manual_AfterMerge.csv', index=True, index_label="@id")


# Option Y: Create merged dataframe from all aggregated dataframes automatically
# Set an empty array to store dataframes
allagg = []
#print(aggregate)
#print("+++++++++++++++++++")
# If there are only 2 dataframes to merge
if len(aggregate) == 2:
    print("Only 2 dataframes+++++++++++++++++++")
    # Indicator causes duplicate column problem from 2nd merge, thus deactivated
    # merged = aggregate[0].merge(aggregate[1].drop_duplicates(), on='full_coverage', how='outer', indicator=True, validate='many_to_many')
    merged = aggregate[0].merge(aggregate[1].drop_duplicates(), on='full_coverage', how='outer', validate='many_to_many')
    allagg.append(merged)
    #print(allagg)
# If there are more than 3 dataframes to merge, loop over dataframes in the aggregate array (created above) until it there is no more dataframes to merge
# Put all dataframes in a new allagg array
else:
    for ii in range(len(aggregate)):
        if ii == 0:
            print(str(ii) + "st merge/n+++++++++++++++++++")
            # Indicator causes duplicate column problem from 2nd merge, thus deactivated
            #merged = aggregate[0].merge(aggregate[1].drop_duplicates(), on='full_coverage', how='outer', indicator=True, validate='many_to_many')
            merged = aggregate[ii].merge(aggregate[ii+1].drop_duplicates(), on='full_coverage', how='outer', validate='many_to_many')
            #merged = aggregate[ii].merge(aggregate[ii+1].drop_duplicates(), on=aggregate[ii].index, how='outer', validate='many_to_many')
            allagg.append(merged)
            #print(allagg)
        elif ii <= len(aggregate)-2:
            print(str(ii) + "th merge/n+++++++++++++++++++")
            # print(len(allagg))
            merged = allagg[ii-1].merge(aggregate[ii+1].drop_duplicates(), on='full_coverage', how='outer', validate='many_to_many')
            allagg.append(merged)
            #print(allagg)

#print("+++++++++++++++++++")
#print(allagg)

# Generate percentage of coverage for each entity
print("COUNTS+++++++++++++++++++")
count = allagg[-1].count(0)
print(count)
#Change series into dataframe
allagg[-1].count().to_frame()

# Save as CSV for the last dafatrame (all merged one) in allagg aray
#allagg_csv = allagg[-1].to_csv('allagg_shakespare.csv', index=True, index_label="@id")
# Save as EXCEL for the last dafatrame (all merged one) in allagg aray
wb = Workbook()
# Access the default sheet and rename it
ws = wb.active
ws.title = "comparison"
# Convert dataframe to rows in EXCEL and add them to the sheet
for r in dataframe_to_rows(allagg[-1], index=True, header=True):
    ws.append(r)
# Alternative way to save as EXCEL
#allagg[-1].to_excel(r'allagg_shakespare.xlsx', sheet_name='comparison', index = False)

print("COUNTS+++++++++++++++++++")
# Obtain http as name for next step
#df['entity'] = df.index
# print(str(str(count[0] / count['full_coverage'] * 100) + '%'))
# print(str(count[2] / count['full_coverage'] * 100) + '%')
# print(str(count[3] / count['full_coverage'] * 100) + '%')

# Create DataFrame for coverage stats
coverage = {'source': [], 'count': [], 'percentage': []}
coverage_df = pd.DataFrame(coverage)
# Iterate over count
for n in range(len(count)):
    rate = str(count[n] / count['full_coverage'] * 100) + '%'
    print(rate)
    coverage_df.loc[n] = [count.index[n], count[n], rate]
    # if n == 1:
    #     continue
    # else:
    #     rate = str(count[n] / count['full_coverage'] * 100) + '%'
    #     print(rate)
    #     coverage_df.loc[n] = [count.index[n], count[n], rate]
print(coverage_df)

# Create a sheet at the end of the workbook
ws1 = wb.create_sheet("coverage")
# Convert dataframe to rows in EXCEL and add them to the sheet
for r in dataframe_to_rows(coverage_df, index=False, header=False):
    ws1.append(r)
# Save Workbook (2 sheets) as XLSX
wb.save('allagg_cesare.xlsx')

# #Saving as CSV file
#df_csv = x.to_csv('testrdf3.csv', index = True, index_label = "@id")
# #Saving as Pickle file
# with open('testrdf.pkl', 'wb') as f:
#     pickle.dump(df, f)
# #Reading Pickle file
# with open('testrdf.pkl', 'rb') as f:
#     df = pickle.load(f)
# df['entity'] = df.index
# print(df.head)
#
# x = df[df['entity'] == "http://dbpedia.org/resource/1967"]
# print(x)
# # #Saving as CSV file
# df_csv = x.to_csv('testrdf2.csv', index = True, index_label = "@id")