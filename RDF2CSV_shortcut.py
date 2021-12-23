import pickle
import rdfpandas
from rdfpandas.graph import to_dataframe
import pandas as pd
import rdflib
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout

# Read EXCEL and put in Dataframe
wb = load_workbook('Persons_Napoleon.xlsx')
#print(wb.sheetnames)
ws = wb['coverage']
#print(ws['A1'].value)
#print(ws['A1':'C9'])

# Create dataframe from EXCEL sheet (iterate each rows in EXCEL)
df = pd.DataFrame(ws.values)
#print(df)
#df.loc[["row1", "row2"],["column1", "column2"]]
for n in range(len(df.index)):
    df.loc[[n],[3]] = df.loc[[n],[2]]
    # Calculate the coverage (prop is the full coverage)
    df.loc[[n],[2]] = df.loc[1,1] - df.loc[n,1]
print(df)

# Create new EXCEL Workbook and sheet
wb1 = Workbook()
ws1 = wb1.create_sheet()
# Create sheet from dataframe (iterate each rows in EXCEL)
for r in dataframe_to_rows(df, index=True, header=True):
    ws1.append(r)
#print(ws1['B3'].value)
#print(ws1['D11'].value)

chart1 = BarChart()
label = Reference(ws1, min_col=2, min_row=3, max_row=11)
data_range = Reference(ws1, min_col=3, max_col=4, min_row=2, max_row=11)
chart1.add_data(data_range, titles_from_data=True)
chart1.set_categories(label)
#chart1.shape = 4
#chart1.type = "col"
chart1.title = "Bar Chart"
chart1.y_axis.title = 'Coverage'
chart1.x_axis.title = 'Source'
chart1.type = "bar"
chart1.style = 13
chart1.grouping = "percentStacked"
chart1.overlap = 100
chart1.title = 'Content Coverage of Sources'
chart1.layout=Layout(
    manualLayout=ManualLayout(
        x=0.2, y=0.2,
        h=0.7, w=1.0,
    )
)
# Specify the starting point of chart
ws1.add_chart(chart1, "F1")


#values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
# chart = BarChart()
# chart.add_data(values)
#
# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Bar Chart"
# chart1.y_axis.title = 'Coverage'
# chart1.x_axis.title = 'Source'
#
# coverage = Reference(ws1, min_col=3, min_row=3, max_row=11)
# full = Reference(ws1, min_col=5, min_row=3, max_row=11)
# chart1.add_data(coverage, titles_from_data=True)
# chart1.set_categories(full)
# chart1.shape = 4
# ws1.add_chart(chart1, "A10")
#
# from copy import deepcopy
# chart4 = deepcopy(chart1)
# chart4.type = "bar"
# chart4.style = 13
# chart4.grouping = "percentStacked"
# chart4.overlap = 100
# chart4.title = 'Percent Stacked Chart'
# ws1.add_chart(chart4, "E15")

# Save wb sheet in Webbook (wb) as xlsx file
wb1.save("Person_Napoleon_chart2.xlsx")

# Load A1:C9 as separate series
#for row in sheet_ranges.iter_rows(min_row=1, max_col=3, max_row=9, values_only=True):
#    print(row)

# # Create the last sheet of the workbook
# ws1 = wb.create_sheet("Copy4")
# # Convert dataframe to rows in EXCEL and add them to the sheet
# for r in dataframe_to_rows(df, index=False, header=False):
#     ws1.append(r)
# # Save as XLSX
# wb.save('allagg_shakespare.xlsx')

# Save in xlsx
#df = Workbook()
#df.save('allagg_shakespare_2sheets.xlsx')

# #print("+++++++++++++++++++")
# #print(allagg)
# # Save as CSV for the last dafatrame (all merged one) in allagg aray
# #allagg_csv = allagg[-1].to_csv('allagg_shakespare.csv', index=True, index_label="@id")
# allagg[-1].to_excel(r'allagg_shakespare.xlsx', sheet_name='comparison', index = False)
#
# # Generate percentage of coverage for each entity
# print("COUNTS+++++++++++++++++++")
# count = allagg[-1].count(0)
# print(count)
# #Change series into dataframe
# allagg[-1].count().to_frame()
#
# print("COUNTS+++++++++++++++++++")
# # Obtain http as name for next step
# #df['entity'] = df.index
# # print(str(str(count[0] / count['full_coverage'] * 100) + '%'))
# # print(str(count[2] / count['full_coverage'] * 100) + '%')
# # print(str(count[3] / count['full_coverage'] * 100) + '%')
#
# # Create DataFrame for coverage stats
# coverage = {'source': [], 'count': [], 'percentage': []}
# coverage_df = pd.DataFrame(coverage)
# # Iterate over count
# for n in range(len(count)):
#     if n == 0:
#         rate = count.index[n] + ': ' + str(str(count[n] / count['full_coverage'] * 100) + '%')
#         print(rate)
#         coverage_df.loc[n] = [count.index[n], count[n], rate]
#     elif n == 1:
#         continue
#     else:
#         print(count.index[n] + ': ' + str(count[n]/count['full_coverage']*100) + '%')
#         coverage_df.loc[n] = [count.index[n], count[n], rate]
# print(coverage_df)
# # Also investigate how to save the allagg in a sheet in the EXCEL above
# coverage_df.to_excel(r'allagg_shakespare.xlsx', sheet_name='coverage', index = False)
